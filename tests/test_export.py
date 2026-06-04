from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from dcf_studio.capm import CAPMInputs, calculate_capm
from dcf_studio.data import CompanyOverview
from dcf_studio.dcf import DCFInputs, calculate_dcf
from dcf_studio.export import export_valuation_workbook
from dcf_studio.wacc import WACCInputs, calculate_wacc


def test_export_workbook_contains_model_sheets_and_formatting():
    overview = CompanyOverview(
        ticker="TEST",
        name="Test Company",
        sector="Technology",
        industry="Software",
        summary="",
        market_cap=1_000.0,
        current_price=100.0,
        enterprise_value=950.0,
        beta=1.1,
        shares_outstanding=10.0,
        cash=50.0,
        debt=100.0,
    )
    historical = pd.DataFrame({"Revenue": [900.0, 1000.0], "Free Cash Flow": [90.0, 120.0]}, index=[2023, 2024])
    forecast = pd.DataFrame({"Free Cash Flow": [130.0, 140.0, 150.0]}, index=[2025, 2026, 2027])
    capm = calculate_capm(CAPMInputs(0.04, 1.1, 0.055))
    wacc = calculate_wacc(WACCInputs(1_000.0, 100.0, capm.cost_of_equity, 0.05, 0.21))
    dcf = calculate_dcf(DCFInputs(forecast, wacc.wacc, 0.025, 50.0, 100.0, 10.0))
    sensitivity = pd.DataFrame({0.02: [90.0], 0.03: [100.0]}, index=[0.09])
    scenarios = pd.DataFrame(
        {
            "intrinsic_value_per_share": [90.0, 100.0, 120.0],
            "probability": [0.25, 0.50, 0.25],
        },
        index=["Bear Case", "Base Case", "Bull Case"],
    )

    workbook_bytes = export_valuation_workbook(
        overview=overview,
        historical=historical,
        forecast=forecast,
        capm=capm,
        wacc=wacc,
        dcf=dcf,
        sensitivity=sensitivity,
        scenarios=scenarios,
    )
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook.sheetnames == [
        "Overview",
        "Historical Financials",
        "Forecast",
        "CAPM",
        "WACC",
        "DCF",
        "Sensitivity",
        "Scenarios",
        "Summary",
    ]
    assert workbook["Summary"].freeze_panes == "A2"
    assert workbook["DCF"]["A1"].font.bold
