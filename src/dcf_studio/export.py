"""Export helpers for model outputs."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from dcf_studio.capm import CAPMResult
from dcf_studio.data import CompanyOverview
from dcf_studio.dcf import DCFResult
from dcf_studio.forecasting import ForecastAssumptions
from dcf_studio.utils import format_currency, format_percent, format_price
from dcf_studio.valuation import upside_downside, valuation_verdict
from dcf_studio.wacc import WACCResult


def build_investment_summary(
    overview: CompanyOverview,
    assumptions: ForecastAssumptions,
    capm: CAPMResult,
    wacc: WACCResult,
    dcf: DCFResult,
) -> str:
    """Generate neutral, README-ready investment summary text."""
    market_price = overview.current_price
    gap = upside_downside(dcf.intrinsic_value_per_share, market_price)
    verdict = valuation_verdict(dcf.intrinsic_value_per_share, market_price)
    return (
        f"{overview.name} ({overview.ticker}) was valued using a {assumptions.horizon_years}-year DCF forecast. "
        f"The model estimates intrinsic value per share of {format_price(dcf.intrinsic_value_per_share)} versus "
        f"a market price of {format_price(market_price)}, implying upside/downside of {format_percent(gap) if gap is not None else 'N/A'}. "
        f"Core assumptions include WACC of {format_percent(wacc.wacc)}, terminal growth of "
        f"{format_percent(dcf.terminal_growth_rate)}, near-term revenue growth of "
        f"{format_percent(assumptions.near_term_revenue_growth)}, and target FCF margin of "
        f"{format_percent(assumptions.target_fcf_margin)}. Based on selected assumptions, the model-implied result is "
        f"{verdict.lower()}. This is an educational valuation exercise, not investment advice or a price target."
    )


def export_valuation_workbook(
    *,
    overview: CompanyOverview,
    historical: pd.DataFrame,
    forecast: pd.DataFrame,
    capm: CAPMResult,
    wacc: WACCResult,
    dcf: DCFResult,
    sensitivity: pd.DataFrame,
    scenarios: pd.DataFrame,
) -> bytes:
    """Export the valuation model to a formatted multi-sheet Excel workbook."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([overview.to_dict()]).T.rename(columns={0: "Value"}).to_excel(writer, sheet_name="Overview")
        historical.to_excel(writer, sheet_name="Historical Financials")
        forecast.to_excel(writer, sheet_name="Forecast")
        _capm_table(capm).to_excel(writer, sheet_name="CAPM", index=False)
        _wacc_table(wacc).to_excel(writer, sheet_name="WACC", index=False)
        _dcf_table(dcf).to_excel(writer, sheet_name="DCF", index=False)
        sensitivity.to_excel(writer, sheet_name="Sensitivity")
        scenarios.to_excel(writer, sheet_name="Scenarios")
        summary = pd.DataFrame(
            {
                "Metric": [
                    "Company",
                    "Ticker",
                    "Market Capitalization",
                    "Enterprise Value",
                    "Equity Value",
                    "Intrinsic Value / Share",
                    "Current Market Price",
                    "Cost of Equity",
                    "WACC",
                    "Terminal Growth",
                ],
                "Value": [
                    overview.name,
                    overview.ticker,
                    format_currency(overview.market_cap),
                    format_currency(dcf.enterprise_value),
                    format_currency(dcf.equity_value),
                    format_price(dcf.intrinsic_value_per_share),
                    format_price(overview.current_price),
                    format_percent(capm.cost_of_equity),
                    format_percent(wacc.wacc),
                    format_percent(dcf.terminal_growth_rate),
                ],
            }
        )
        summary.to_excel(writer, sheet_name="Summary", index=False)

        _format_workbook(writer.book)

    return output.getvalue()


def _capm_table(capm: CAPMResult) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Input": ["Risk-Free Rate", "Beta", "Market Risk Premium", "Cost of Equity"],
            "Value": [capm.risk_free_rate, capm.beta, capm.market_risk_premium, capm.cost_of_equity],
        }
    )


def _wacc_table(wacc: WACCResult) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Input": [
                "Equity Weight",
                "Debt Weight",
                "Cost of Equity",
                "After-Tax Cost of Debt",
                "Tax Rate",
                "WACC",
            ],
            "Value": [
                wacc.equity_weight,
                wacc.debt_weight,
                wacc.cost_of_equity,
                wacc.after_tax_cost_of_debt,
                wacc.tax_rate,
                wacc.wacc,
            ],
        }
    )


def _dcf_table(dcf: DCFResult) -> pd.DataFrame:
    bridge = pd.DataFrame(
        {
            "Metric": [
                "PV of Forecast Free Cash Flows",
                "Terminal Value",
                "PV of Terminal Value",
                "Enterprise Value",
                "Less: Debt",
                "Add: Cash",
                "Equity Value",
                "Shares Outstanding",
                "Intrinsic Value / Share",
            ],
            "Value": [
                dcf.pv_of_forecast_cash_flows,
                dcf.terminal_value,
                dcf.pv_terminal_value,
                dcf.enterprise_value,
                dcf.debt,
                dcf.cash,
                dcf.equity_value,
                dcf.shares_outstanding,
                dcf.intrinsic_value_per_share,
            ],
        }
    )
    return pd.concat(
        [
            pd.DataFrame(
                {"Metric": ["Discount Rate", "Terminal Growth"], "Value": [dcf.discount_rate, dcf.terminal_growth_rate]}
            ),
            bridge,
        ],
        ignore_index=True,
    )


def _format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    percent_keywords = ("margin", "growth", "rate", "wacc", "weight", "probability", "tax", "cost of equity")
    currency_keywords = (
        "revenue",
        "ebitda",
        "income",
        "cash flow",
        "value",
        "cash",
        "debt",
        "capitalization",
        "price",
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        headers = {
            column_index: str(worksheet.cell(row=1, column=column_index).value or "").lower()
            for column_index in range(1, worksheet.max_column + 1)
        }
        for row in worksheet.iter_rows(min_row=2):
            row_label = " ".join(str(cell.value or "").lower() for cell in row[:2])
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    header = headers.get(cell.column, "")
                    label_context = f"{header} {row_label}"
                    if any(keyword in label_context for keyword in percent_keywords):
                        cell.number_format = "0.0%"
                    elif any(keyword in label_context for keyword in currency_keywords):
                        cell.number_format = "$#,##0.00;[Red]($#,##0.00)"
                    elif "shares" in label_context:
                        cell.number_format = "#,##0"

        for column_index in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            max_length = max(
                len(str(worksheet.cell(row=row_index, column=column_index).value or ""))
                for row_index in range(1, worksheet.max_row + 1)
            )
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 36)
